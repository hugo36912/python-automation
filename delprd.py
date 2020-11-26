import os
from datetime import datetime
from pathlib import Path
import sys
from openpyxl import load_workbook

def grpvoke(name,config,ip):
    sour = 0
    dest = 0
    for tc in config:
      if "policy" in tc and "from-zone" in tc:
        cur = tc.split()
        if cur[8]==name and cur[10] == "source-address" :
            if cur[11] not in ip :
                sour += 1
        if cur[8]==name and cur[10] == "destination-address" :
            if cur[11] not in ip :
                dest += 1
    if dest == 0 or sour == 0:
        f.write("###delete policy of gp###\n")
        g.write("###delete policy of gp###\n")
        for cc in config:
            if "policy" in cc and "from-zone" in cc:
                tc = cc.split()
                if tc[8] == name:
                    cc = cc.replace("set", "delete")
                    f.write(cc)
                    g.write(cc)
    else:
        f.write("###delete address of gp ###\n")
        g.write("###delete address of gp ###\n")
        for ct in config:
            for iy in ip:
              if "policy" in ct and "from-zone" in ct:
                tc = ct.split()
                if name == tc[8] and iy in ct:
                    ct = ct.replace("set", "delete")
                    f.write(ct)
                    g.write(ct)
def policyvoke(name,config,ip):
    sour=0
    dest=0
    for cur in config:
        if "policy" in cur and "from-zone" in cur:
            cur=cur.split()
            if cur[8]==name and cur[10] == "source-address" :
                if cur[11].split('-')[0]+'-' not in ip:
                  sour+=1
            if cur[8]==name  and cur[10] == "destination-address" :
                if cur[11].split('-')[0]+'-' not in ip:
                  dest+=1
    if dest == 0 or sour == 0:
        f.write("###delete policy###\n")
        g.write("###delete policy###\n")
        for cc in config:
            if "policy" in cc and "from-zone" in cc:
             tc=cc.split()
             if tc[8]==name:
               cc=cc.replace("set","delete")
               f.write(cc)
               g.write(cc)
    else :
        f.write("###delete address###\n")
        for ct in config:
            for iy in ip:
             if "policy" in ct and "from-zone" in ct:
              tc=ct.split()
              if name ==tc[8] and iy in ct:
                ct=ct.replace("set", "delete")
                f.write(ct)
                g.write(ct)
def getcon(pa):
    deviceFullPathName = "Z:\\Network_Security\\Firewall_Policy_Program\\PythonFirewall\\PRDv3\\tmpConfigFile\\"+pa
    devv = open(deviceFullPathName, "r")
    con=devv.readlines()
     for curReadStr in con:
                if "#" in curReadStr:
                    continue
                if "delete" in curReadStr and "schedulers" not in curReadStr :
                    hj=curReadStr
                    hj=hj.replace("delete","set")
                    if hj in DeviceConfig:
                     DeviceConfig.remove(hj)
                     continue
                if "from-zone" in curReadStr and "to-zone" in curReadStr:
                    DeviceConfig.append(curReadStr)
                if "address-book" in curReadStr and "address-set" in curReadStr:
                    DeviceConfig.append(curReadStr)
                if "address-book" in curReadStr and "address-set" not in curReadStr:   
                    DeviceConfig.append(curReadStr)
                if "set applications application" in curReadStr:
                    DeviceConfig.append(curReadStr) 
        return DeviceConfig   
des=input("description:")
num=input("num:")
inputip=[]
Path("Z:\\Network_Security\\Firewall_Policy_Program\\PythonFirewall\\PRDv3\\TSRcmd\\"+des).mkdir(parents=True, exist_ok=True)
inp=open('Z:\\Network_Security\\Firewall_Policy_Program\\PythonFirewall\\PRDv3\\TSRcmd\\'+des+'\\input.txt','a')
if num.isnumeric()==False:
       ip=num.split('/', 3)
       if len(ip)<3:
           ip[0] = ip[0].replace('-', '')
           ip[0] = ip[0].replace('- ', '')
           ip[0] = ip[0].replace(' ', '')
           inputip.append(ip[0]+'-')
           inp.write('delete ip: ' + ip[0]+"\n")
       else:
           ip[0]=ip[0].replace('-','')
           ip[0]=ip[0].replace('- ', '')
           ip[0]=ip[0].replace(' ', '')
           ip[1]=ip[1].replace(' ','')
           ip[2]=ip[2].replace(' ','')
           inputip.append(ip[0]+'-')
           inp.write('delete ip: '+ip[0]+" hostname: "+ip[1]+"\n")
           num=1
else:
     for i in range(int(num)):
      ip=input().split('/', 3)
      if len(ip) < 3:
          ip[0] = ip[0].replace('-', '')
          ip[0] = ip[0].replace('- ', '')
          ip[0] = ip[0].replace(' ', '')
          inputip.append(ip[0]+'-')
          inp.write('delete ip: ' + ip[0]+"\n")
      else:
          ip[0]=ip[0].replace('-','')
          ip[0]=ip[0].replace('- ', '')
          ip[0]=ip[0].replace(' ', '')
          ip[1]=ip[1].replace(' ','')
          ip[2]=ip[2].replace(' ','')
          inputip.append(ip[0]+'-')
          inp.write('delete ip: '+ip[0]+" hostname: "+ip[1]+"\n")


gp=['DC1_eHR_FW_C01','DC1_eHR_FW_D01','DC1_eHR_FW_E01','DC1_eHR_FW_F01','DC2_eHR_FW_C01','DC2_eHR_FW_D01','DC2_eHR_FW_E01'\
,'DC2_eHR_FW_F01','DC6_PUB_FW_C01','DC6_PUB_FW_D01','DC6_PUB_FW_E01','DC6_PUB_FW_F01','DC7_PUB_FW_C01','DC7_PUB_FW_D01',\
'DC7_PUB_FW_E01','DC7_PUB_FW_F01']

    
for g in gp:
    print("loading")
    contents = getcon()
    f=open('Z:\\Network_Security\\Firewall_Policy_Program\\PythonFirewall\\PRDv3\\TSRcmd\\'+des+"\\"+g,'a')
    g=open("Z:\\Network_Security\\Firewall_Policy_Program\\PythonFirewall\\PRDv3\\tmpConfigFile\\"+g,'a')
    name=[]
    gpname=[]
    addrset=[]
    finaddr=[]
    for curReadStr in contents:
     if "#" in curReadStr:
         continue
     if "address-set" in curReadStr and curReadStr.split()[7] in addrset:
         addrset.append(curReadStr.split()[7])
     for iq in inputip:
      if iq in curReadStr and "address-book" in curReadStr:
        curReadStr=curReadStr.replace("set","delete")
        curReadStr = curReadStr.replace("address-delete", "address-set")
        if "address-set" in curReadStr :
            adrg = curReadStr.split()
            addrset.append(adrg[7])
        f.write(curReadStr)
        g.write(curReadStr)
      if iq in curReadStr and "policy" in curReadStr and "from-zone" in curReadStr:
         poci=curReadStr.split()
         if poci[8] not in name:
          name.append(poci[8])
    for na in name :
       policyvoke(na,contents,inputip)
    for cr in addrset:
        if addrset.count(cr)==1:
            finaddr.append(cr)
            for ck in contents:
                if "#" in ck:
                    continue
                if  cr in ck and "policy" in ck and "from-zone" in ck:
                    poge = ck.split()
                    if poge[8] not in gpname:
                        gpname.append(poge[8])
    for na in gpname:
     grpvoke(na, contents, finaddr)
wb = load_workbook("Z:/Network_Security/Firewall_Policy_Program/PythonFirewall/PRDv3/connectDatabase/Database/eHR_PRD_FW_gpinput.xlsx")
ws=wb['group_lookup']
max=ws.max_row
for k in range(1,max+1):
    if ws.cell(k,3).value+'-' in [ele for ele in inputip] :
        ws.delete_rows(k)
        print("delete:",ws.cell(k,3).value)
wb.save("Z:/Network_Security/Firewall_Policy_Program/PythonFirewall/PRDv3/connectDatabase/Database/eHR_PRD_FW_gpinput.xlsx")
print("success")
