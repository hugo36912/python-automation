import os
from openpyxl import load_workbook
import pyperclip
def jj():
   inputgrp=input("groupname:")
   num=input("num:")
   inputip=[]
   go=True
   f =open('Z:\\Network_Security\Firewall_Policy_Program\PythonFirewall\PRDv3\TSRcmd\cptable.txt','a')
   f.write("group:"+inputgrp+'\n')
   os=''
   osj=''
   osl = inputgrp.split('-')
   if len(osl)>2:
    os = osl[3]
    osj=osl[2]
   if os =='LNX':
       os="LINUX"
   if num.isnumeric()==False:
       ip=num.split('/', 3)
       ip[0]=ip[0].replace('-','')
       ip[0]=ip[0].replace('- ', '')
       ip[0]=ip[0].replace(' ', '')
       ip[1]=ip[1].replace(' ','')
       ip[2]=ip[2].replace(' ','')
       inputip.append(ip[0]+'-'+ip[1])
       f.write("S2"+" "+os+" "+inputgrp+' '+ip[0]+" "+ip[1]+" "+osj+"\n")
       num=1
   else:
     for i in range(int(num)):
      ip=input().split('/', 3)
      ip[0]=ip[0].replace('-','')
      ip[0]=ip[0].replace('- ', '')
      ip[0]=ip[0].replace(' ', '')
      ip[1]=ip[1].replace(' ','')
      ip[2]=ip[2].replace(' ','')
      inputip.append(ip[0]+'-'+ip[1])
      f.write("S2"+" "+os+" "+inputgrp+' '+ip[0]+" "+ip[1]+" "+osj+"\n")

   wb = load_workbook("Z:/Network_Security/Firewall_Policy_Program/PythonFirewall/PRDv3/connectDatabase/Database/eHR_PRD_FW_gpinput.xlsx")
   ws=wb['group_lookup']
   max=ws.max_row
   if ws.cell(max,1).value ==None and ws.cell(max,2).value==None and ws.cell(max,3).value ==None and ws.cell(max,4).value ==None :
        newr=max
   else:
     newr=max+1
   for c in range(1,newr):
     if inputgrp == ws.cell(c,1).value and ws.cell(c,2).value in inputip:
       print('gp exist',"gp name:"+ws.cell(c,1).value,"host:"+ws.cell(c,2).value,"ip:"+ws.cell(c,3).value)
       go=False
     if (ws.cell(c,2).value in inputip or ws.cell(c,3).value in inputip) and inputgrp != ws.cell(c,1).value:
       print('check excel',"gp name:"+ws.cell(c,1).value,"host:"+ws.cell(c,2).value,"ip:"+ws.cell(c,3).value)
   if go == True:
     for k in range(int(num)):
      ws.cell(newr,1).value=inputgrp
      ws.cell(newr,2).value=inputip[k]
      ws.cell(newr,3).value=inputip[k].split('-',1)[0]
      ws.cell(newr,4).value=32
      newr+=1
   wb.save("Z:/Network_Security/Firewall_Policy_Program/PythonFirewall/PRDv3/connectDatabase/Database/eHR_PRD_FW_gpinput.xlsx")
   print("success")
   pyperclip.copy(inputgrp)
   f.write("---------------------------------------\n")
   print("copied gpname to clipboard and stord in note")
while(True):
  jj()